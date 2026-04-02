"""Plot-mode data profiling helpers extracted from openplot.server."""

from __future__ import annotations

from pathlib import Path
from types import ModuleType
from typing import cast

import pandas as pd
from openpyxl import load_workbook

from .models import (
    PlotModeDataProfile,
    PlotModeDataRegion,
    PlotModeFile,
    PlotModeSheetBounds,
    PlotModeSheetCandidate,
    PlotModeSheetPreview,
    PlotModeTabularSelector,
)


def _stringify_preview_value(server_module: ModuleType, value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value)
    if len(text) > 120:
        return text[:117] + "..."
    return text


def _sample_integrity_notes(
    server_module: ModuleType, frame: pd.DataFrame
) -> list[str]:
    notes: list[str] = []
    if frame.empty:
        notes.append(
            "The sampled rows are empty, so I need confirmation before drafting a plot."
        )
        return notes

    empty_columns: list[str] = []
    for column in frame.columns:
        series = cast(pd.Series, frame[column])
        if bool(series.isna().all()):
            empty_columns.append(str(column))
    if empty_columns:
        notes.append(
            "The sample includes all-empty columns: " + ", ".join(empty_columns[:5])
        )

    blank_row_count = int(frame.isna().all(axis=1).sum())
    if blank_row_count:
        notes.append(f"The sample includes {blank_row_count} fully empty row(s).")

    duplicate_columns = frame.columns[frame.columns.duplicated()].tolist()
    if duplicate_columns:
        notes.append(
            "Duplicate column labels detected: "
            + ", ".join(str(value) for value in duplicate_columns[:5])
        )

    missing_columns: list[str] = []
    for column in frame.columns:
        series = cast(pd.Series, frame[column])
        if len(series.index) == 0:
            continue
        missing_ratio = float(cast(float, series.isna().mean()))
        if missing_ratio >= 0.5:
            missing_columns.append(f"{column} ({missing_ratio:.0%} missing in sample)")
    if missing_columns:
        notes.append(
            "High missingness in sampled columns: " + ", ".join(missing_columns[:5])
        )

    return notes


def _column_label(server_module: ModuleType, index: int) -> str:
    label = ""
    value = index + 1
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        label = chr(65 + remainder) + label
    return label or "A"


def _format_sheet_bounds(
    server_module: ModuleType, bounds: tuple[int, int, int, int]
) -> str:
    row_start, row_end, col_start, col_end = bounds
    return (
        f"{server_module._column_label(col_start)}{row_start + 1}:"
        f"{server_module._column_label(col_end)}{row_end + 1}"
    )


def _format_sheet_region_label(
    server_module: ModuleType,
    sheet_name: str | None,
    bounds: tuple[int, int, int, int] | None,
) -> str:
    if bounds is None:
        return sheet_name or "<unknown region>"
    bounds_label = server_module._format_sheet_bounds(bounds)
    if sheet_name:
        return f"{sheet_name}!{bounds_label}"
    return bounds_label


def _normalize_preview_grid(
    server_module: ModuleType, rows: list[list[str]]
) -> list[list[str]]:
    width = max((len(row) for row in rows), default=0)
    if width <= 0:
        return []
    return [row + [""] * (width - len(row)) for row in rows]


def _non_empty_cell_count(server_module: ModuleType, row: list[str]) -> int:
    return sum(1 for value in row if value.strip())


def _detect_non_empty_blocks(
    server_module: ModuleType, rows: list[list[str]]
) -> list[tuple[int, int, int, int]]:
    grid = server_module._normalize_preview_grid(rows)
    if not grid:
        return []

    height = len(grid)
    width = len(grid[0])
    visited = [[False for _ in range(width)] for _ in range(height)]
    bounds_list: list[tuple[int, int, int, int]] = []

    for row_index in range(height):
        for col_index in range(width):
            if visited[row_index][col_index] or not grid[row_index][col_index].strip():
                continue

            queue = [(row_index, col_index)]
            visited[row_index][col_index] = True
            cells: list[tuple[int, int]] = []

            while queue:
                current_row, current_col = queue.pop()
                cells.append((current_row, current_col))
                for next_row, next_col in (
                    (current_row - 1, current_col),
                    (current_row + 1, current_col),
                    (current_row, current_col - 1),
                    (current_row, current_col + 1),
                ):
                    if not (0 <= next_row < height and 0 <= next_col < width):
                        continue
                    if visited[next_row][next_col]:
                        continue
                    if not grid[next_row][next_col].strip():
                        continue
                    visited[next_row][next_col] = True
                    queue.append((next_row, next_col))

            if not cells:
                continue
            min_row = min(row for row, _ in cells)
            max_row = max(row for row, _ in cells)
            min_col = min(col for _, col in cells)
            max_col = max(col for _, col in cells)
            if len(cells) >= 2:
                bounds_list.append((min_row, max_row, min_col, max_col))

    if bounds_list:
        return sorted(bounds_list)

    non_empty_rows = [
        index
        for index, row in enumerate(grid)
        if server_module._non_empty_cell_count(row)
    ]
    non_empty_cols = [
        index
        for index in range(width)
        if any(grid[row_index][index].strip() for row_index in range(height))
    ]
    if not non_empty_rows or not non_empty_cols:
        return []
    return [
        (
            non_empty_rows[0],
            non_empty_rows[-1],
            non_empty_cols[0],
            non_empty_cols[-1],
        )
    ]


def _rows_for_bounds(
    server_module: ModuleType,
    rows: list[list[str]],
    bounds: tuple[int, int, int, int],
) -> list[list[str]]:
    row_start, row_end, col_start, col_end = bounds
    normalized = server_module._normalize_preview_grid(rows)
    sliced = [
        row[col_start : col_end + 1] for row in normalized[row_start : row_end + 1]
    ]
    while sliced and not any(cell.strip() for cell in sliced[-1]):
        sliced.pop()
    while (
        sliced
        and server_module._non_empty_cell_count(sliced[0]) <= 1
        and len(sliced) > 1
    ):
        sliced = sliced[1:]
    return sliced


def _looks_like_numeric_text(server_module: ModuleType, value: str) -> bool:
    stripped = value.strip().replace(",", "")
    if not stripped:
        return False
    if stripped.endswith("%"):
        stripped = stripped[:-1]
    try:
        float(stripped)
    except ValueError:
        return False
    return True


def _dataframe_from_block_rows(
    server_module: ModuleType, rows: list[list[str]]
) -> pd.DataFrame:
    normalized = server_module._normalize_preview_grid(rows)
    if not normalized:
        return pd.DataFrame()

    first_row = normalized[0]
    non_empty_first_row = [value for value in first_row if value.strip()]
    use_header = bool(non_empty_first_row) and (
        any(any(char.isalpha() for char in value) for value in non_empty_first_row)
        or not all(
            server_module._looks_like_numeric_text(value)
            for value in non_empty_first_row
        )
    )
    if use_header:
        headers: list[str] = []
        seen_headers: set[str] = set()
        for index, value in enumerate(first_row):
            candidate = value.strip() or f"column_{index + 1}"
            deduped = candidate
            suffix = 2
            while deduped in seen_headers:
                deduped = f"{candidate}_{suffix}"
                suffix += 1
            headers.append(deduped)
            seen_headers.add(deduped)
        data_rows = normalized[1:]
        return pd.DataFrame(data_rows, columns=headers)

    headers = [f"column_{index + 1}" for index in range(len(first_row))]
    return pd.DataFrame(normalized, columns=headers)


def _build_data_profile(
    server_module: ModuleType,
    *,
    file_path: Path,
    file_id: str | None,
    source_kind: str,
    source_label: str,
    table_name: str | None,
    frame: pd.DataFrame,
    inferred_bounds: tuple[int, int, int, int] | None = None,
) -> PlotModeDataProfile:
    normalized_frame = frame.head(8).copy()
    normalized_frame = normalized_frame.replace({pd.NA: None})
    preview_rows = [
        [server_module._stringify_preview_value(value) for value in row]
        for row in normalized_frame.to_numpy().tolist()
    ]
    columns = [str(column) for column in normalized_frame.columns.tolist()]
    integrity_notes = server_module._sample_integrity_notes(normalized_frame)
    summary = f"{source_kind.title()} preview with {len(columns)} sampled column(s)" + (
        f" from {table_name}" if table_name else ""
    )
    return PlotModeDataProfile(
        file_path=str(file_path),
        file_name=file_path.name,
        source_label=source_label,
        source_kind=source_kind,
        table_name=table_name,
        summary=summary,
        columns=columns,
        preview_rows=preview_rows,
        integrity_notes=integrity_notes,
        needs_confirmation=source_kind == "excel" or table_name is not None,
        source_file_id=file_id,
        inferred_sheet_name=table_name,
        inferred_bounds=inferred_bounds,
    )


def _build_tabular_region_from_frame(
    server_module: ModuleType,
    *,
    file_path: Path,
    source_kind: str,
    sheet_name: str | None,
    bounds: tuple[int, int, int, int],
    frame: pd.DataFrame,
) -> PlotModeDataRegion:
    normalized_frame = frame.head(8).copy()
    normalized_frame = normalized_frame.replace({pd.NA: None})
    preview_rows = [
        [server_module._stringify_preview_value(value) for value in row]
        for row in normalized_frame.to_numpy().tolist()
    ]
    columns = [str(column) for column in normalized_frame.columns.tolist()]
    bounds_label = server_module._format_sheet_bounds(bounds)
    source_label = file_path.name
    if sheet_name:
        source_label = f"{file_path.name} - {sheet_name} ({bounds_label})"
    else:
        source_label = f"{file_path.name} ({bounds_label})"
    summary = f"Sampled {source_kind} table from {bounds_label}" + (
        f" on {sheet_name}" if sheet_name else ""
    )
    return PlotModeDataRegion(
        sheet_name=sheet_name,
        source_label=source_label,
        summary=summary,
        bounds=server_module._sheet_bounds_from_tuple(bounds),
        columns=columns,
        preview_rows=preview_rows,
    )


def _build_data_profile_from_grid(
    server_module: ModuleType,
    *,
    file_path: Path,
    file_id: str,
    source_kind: str,
    sheet_name: str | None,
    bounds: tuple[int, int, int, int],
    rows: list[list[str]],
) -> PlotModeDataProfile:
    frame = server_module._dataframe_from_block_rows(
        server_module._rows_for_bounds(rows, bounds)
    )
    tabular_region = server_module._build_tabular_region_from_frame(
        file_path=file_path,
        source_kind=source_kind,
        sheet_name=sheet_name,
        bounds=bounds,
        frame=frame,
    )
    profile = server_module._build_data_profile(
        file_path=file_path,
        file_id=file_id,
        source_kind=source_kind,
        source_label=tabular_region.source_label,
        table_name=sheet_name,
        frame=frame,
        inferred_bounds=bounds,
    )
    profile.summary = tabular_region.summary
    profile.needs_confirmation = True
    profile.tabular_regions = [tabular_region]
    return profile


def _build_grouped_data_profile_from_regions(
    server_module: ModuleType,
    *,
    file_path: Path,
    file_id: str,
    source_kind: str,
    region_profiles: list[PlotModeDataProfile],
) -> PlotModeDataProfile:
    if not region_profiles:
        raise ValueError("At least one region profile is required.")
    if len(region_profiles) == 1:
        return region_profiles[0]

    tabular_regions = [
        region
        for profile in region_profiles
        for region in (profile.tabular_regions or [])
    ]
    if not tabular_regions:
        return region_profiles[0]

    sheet_names = sorted(
        {
            region.sheet_name.strip()
            for region in tabular_regions
            if region.sheet_name and region.sheet_name.strip()
        }
    )
    integrity_notes: list[str] = []
    for profile in region_profiles:
        for note in profile.integrity_notes:
            if note not in integrity_notes:
                integrity_notes.append(note)

    region_count = len(tabular_regions)
    sheet_count = len(sheet_names)
    source_label = f"{file_path.name} - {region_count} selected regions"
    summary = f"Sampled {region_count} {source_kind} regions"
    if sheet_count == 1 and sheet_names:
        summary += f" on {sheet_names[0]}"
    elif sheet_count > 1:
        summary += f" across {sheet_count} sheets"

    return PlotModeDataProfile(
        file_path=str(file_path),
        file_name=file_path.name,
        source_label=source_label,
        source_kind=source_kind,
        summary=summary,
        integrity_notes=integrity_notes,
        needs_confirmation=True,
        source_file_id=file_id,
        tabular_regions=tabular_regions,
    )


def _build_sheet_preview(
    server_module: ModuleType,
    *,
    sheet_name: str,
    rows: list[list[str]],
    total_rows: int,
    total_cols: int,
) -> PlotModeSheetPreview:
    bounds_list = server_module._detect_non_empty_blocks(rows)
    candidates = [
        PlotModeSheetCandidate(
            label=f"Candidate {index + 1} ({server_module._format_sheet_bounds(bounds)})",
            bounds=PlotModeSheetBounds(
                row_start=bounds[0],
                row_end=bounds[1],
                col_start=bounds[2],
                col_end=bounds[3],
            ),
            summary=f"Detected non-empty table block in {sheet_name}",
        )
        for index, bounds in enumerate(bounds_list)
    ]
    return PlotModeSheetPreview(
        name=sheet_name,
        total_rows=total_rows,
        total_cols=total_cols,
        preview_rows=server_module._normalize_preview_grid(rows),
        candidate_tables=candidates,
    )


def _read_delimited_grid(
    server_module: ModuleType,
    path: Path,
    *,
    delimiter: str,
) -> tuple[list[list[str]], int, int]:
    frame = pd.read_csv(
        path,
        sep=delimiter,
        header=None,
        dtype=str,
        keep_default_na=False,
        engine="python",
        on_bad_lines="skip",
    )
    rows = [
        [server_module._stringify_preview_value(value) for value in row]
        for row in frame.to_numpy().tolist()
    ]
    return rows, int(frame.shape[0]), int(frame.shape[1])


def _build_tabular_selector(
    server_module: ModuleType,
    *,
    file: PlotModeFile,
    path: Path,
    source_kind: str,
    sheets: list[PlotModeSheetPreview],
) -> PlotModeTabularSelector:
    return PlotModeTabularSelector(
        file_id=file.id,
        file_path=str(path),
        file_name=path.name,
        source_kind=source_kind,
        sheets=sheets,
        selected_sheet_id=sheets[0].id if sheets else None,
        status_text=(
            f"I found multiple possible tables in {path.name}. Mark one or more regions that belong to the source you want."
        ),
        requires_user_hint=True,
    )


def _tabular_regions_for_profile(
    server_module: ModuleType,
    profile: PlotModeDataProfile,
) -> list[PlotModeDataRegion]:
    if profile.tabular_regions:
        return profile.tabular_regions
    if profile.inferred_bounds is None:
        return []
    return [
        PlotModeDataRegion(
            sheet_name=profile.inferred_sheet_name,
            source_label=profile.source_label,
            summary=profile.summary,
            bounds=server_module._sheet_bounds_from_tuple(profile.inferred_bounds),
            columns=profile.columns,
            preview_rows=profile.preview_rows,
        )
    ]


def _profile_delimited_file(
    server_module: ModuleType,
    file: PlotModeFile,
    path: Path,
    *,
    delimiter: str,
    source_kind: str,
) -> tuple[list[PlotModeDataProfile], PlotModeTabularSelector | None, list[str]]:
    rows, total_rows, total_cols = server_module._read_delimited_grid(
        path, delimiter=delimiter
    )
    sheet_preview = server_module._build_sheet_preview(
        sheet_name=path.name,
        rows=rows,
        total_rows=total_rows,
        total_cols=total_cols,
    )
    if len(sheet_preview.candidate_tables) <= 1:
        if not sheet_preview.candidate_tables:
            return (
                [],
                None,
                [f"Read {path.name}, but did not detect a clear table region."],
            )
        profile = server_module._build_data_profile_from_grid(
            file_path=path,
            file_id=file.id,
            source_kind=source_kind,
            sheet_name=None,
            bounds=server_module._bounds_from_sheet_bounds(
                sheet_preview.candidate_tables[0].bounds
            ),
            rows=sheet_preview.preview_rows,
        )
        return [profile], None, [f"Read {path.name} and found one likely table."]

    selector = server_module._build_tabular_selector(
        file=file,
        path=path,
        source_kind=source_kind,
        sheets=[sheet_preview],
    )
    return (
        [],
        selector,
        [
            f"Read {path.name} and found {len(sheet_preview.candidate_tables)} candidate tables."
        ],
    )


def _profile_json_file(
    server_module: ModuleType, path: Path
) -> list[PlotModeDataProfile]:
    frame = pd.read_json(path, lines=path.suffix.lower() == ".jsonl").head(8)
    return [
        server_module._build_data_profile(
            file_path=path,
            file_id=None,
            source_kind="json",
            source_label=path.name,
            table_name=None,
            frame=frame,
        )
    ]


def _profile_excel_file(
    server_module: ModuleType,
    file: PlotModeFile,
    path: Path,
) -> tuple[list[PlotModeDataProfile], PlotModeTabularSelector | None, list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheets: list[PlotModeSheetPreview] = []
    total_candidates = 0
    non_empty_sheet_count = 0
    try:
        for worksheet in workbook.worksheets[:8]:
            max_row = int(worksheet.max_row or 0)
            max_col = int(worksheet.max_column or 0)
            rows: list[list[str]] = []
            if max_row > 0 and max_col > 0:
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=max_row,
                    min_col=1,
                    max_col=max_col,
                    values_only=True,
                ):
                    rows.append(
                        [server_module._stringify_preview_value(value) for value in row]
                    )
            sheet_preview = server_module._build_sheet_preview(
                sheet_name=worksheet.title,
                rows=rows,
                total_rows=int(worksheet.max_row or len(rows)),
                total_cols=int(worksheet.max_column or (len(rows[0]) if rows else 0)),
            )
            if sheet_preview.candidate_tables:
                non_empty_sheet_count += 1
                total_candidates += len(sheet_preview.candidate_tables)
            sheets.append(sheet_preview)
    finally:
        workbook.close()

    if total_candidates == 1:
        for sheet in sheets:
            if not sheet.candidate_tables:
                continue
            profile = server_module._build_data_profile_from_grid(
                file_path=path,
                file_id=file.id,
                source_kind="excel",
                sheet_name=sheet.name,
                bounds=server_module._bounds_from_sheet_bounds(
                    sheet.candidate_tables[0].bounds
                ),
                rows=sheet.preview_rows,
            )
            return (
                [profile],
                None,
                [f"Read {path.name} and found one likely table on {sheet.name}."],
            )

    visible_sheets = [
        sheet for sheet in sheets if sheet.candidate_tables or sheet.preview_rows
    ]
    if not visible_sheets:
        return (
            [],
            None,
            [f"Read {path.name}, but did not detect a usable worksheet preview."],
        )

    selector = server_module._build_tabular_selector(
        file=file,
        path=path,
        source_kind="excel",
        sheets=visible_sheets,
    )
    return (
        [],
        selector,
        [
            f"Read {path.name} and found {max(total_candidates, non_empty_sheet_count)} possible source tables across workbook sheets."
        ],
    )


def _profile_selected_data_files(
    server_module: ModuleType,
    files: list[PlotModeFile],
) -> tuple[list[PlotModeDataProfile], list[str], PlotModeTabularSelector | None]:
    profiles: list[PlotModeDataProfile] = []
    activity_items: list[str] = []
    selector: PlotModeTabularSelector | None = None
    for file in files:
        path = Path(file.stored_path).resolve()
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                file_profiles, file_selector, file_activity = (
                    server_module._profile_delimited_file(
                        file,
                        path,
                        delimiter=",",
                        source_kind="csv",
                    )
                )
            elif suffix == ".tsv":
                file_profiles, file_selector, file_activity = (
                    server_module._profile_delimited_file(
                        file,
                        path,
                        delimiter="\t",
                        source_kind="tsv",
                    )
                )
            elif suffix in {".json", ".jsonl"}:
                file_profiles = server_module._profile_json_file(path)
                file_selector = None
                file_activity = [
                    f"Read {path.name} and sampled {sum(len(profile.preview_rows) for profile in file_profiles)} row(s)."
                ]
            elif suffix in {".xls", ".xlsx"}:
                file_profiles, file_selector, file_activity = (
                    server_module._profile_excel_file(file, path)
                )
            else:
                file_profiles = []
                file_selector = None
                file_activity = []
        except Exception as exc:
            activity_items.append(
                f"Tried to inspect {path.name}, but preview loading failed: {exc}"
            )
            continue

        activity_items.extend(file_activity)

        if selector is None and file_selector is not None:
            selector = file_selector
        elif file_selector is not None:
            activity_items.append(
                f"{path.name} also needs range selection; resolve one source at a time."
            )

        if file_profiles:
            profiles.extend(file_profiles)
            continue

        if file_selector is not None:
            continue

        activity_items.append(
            f"Registered {path.name}, but automatic preview is unavailable for {suffix or 'this file type'}."
        )
        profiles.append(
            PlotModeDataProfile(
                file_path=str(path),
                file_name=path.name,
                source_label=path.name,
                source_kind="file",
                summary="Unsupported preview type; use the absolute path directly in the plot script.",
                needs_confirmation=True,
                source_file_id=file.id,
            )
        )

    return profiles, activity_items, selector
